import string
import time
import random

import openpyxl
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service as chrome_service
from selenium.webdriver.firefox.service import Service as firefox_service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

f_service = firefox_service('drivers/geckodriver.exe')
opt = webdriver.FirefoxOptions()
driver = webdriver.Firefox(service=f_service)
driver.maximize_window()
driver.implicitly_wait(10)
driver.get("http://automationpractice.com/")

time.sleep(2)
driver.find_element(By.XPATH, "//a[@class='login']").click()


def random_email():
    global email
    email = ''.join(random.choice(string.ascii_letters) for x in range(10)) + "12@gmail.com"
    return email


e = random_email()
print(e)
time.sleep(3)
driver.find_element(By.XPATH, "//input[@name='email_create']").send_keys(random_email())
driver.find_element(By.ID, "SubmitCreate").click()
time.sleep(2)

# Create account

driver.find_element(By.XPATH, "//div[@class='radio-inline'][2]").click()
driver.find_element(By.XPATH, "//input[@data-validate='isName'][1]").send_keys("Mercy")
time.sleep(2)
driver.find_element(By.ID, "customer_lastname").send_keys("Angel")
driver.find_element(By.ID, "email").click()
driver.find_element(By.ID, "passwd").send_keys("abcd1234")
password = driver.find_element(By.ID, "passwd").text

select_date = Select(driver.find_element(By.XPATH, "//select[@id='days']"))
select_date.select_by_index(6)
select_month = Select(driver.find_element(By.XPATH, "//select[@id='months']"))
select_month.select_by_index(8)
select_year = Select(driver.find_element(By.XPATH, "//select[@id='years']"))
select_year.select_by_index(20)
driver.find_element(By.ID, "firstname").click()
driver.find_element(By.ID, "lastname").click()
driver.find_element(By.ID, "company").send_keys('Random')
driver.find_element(By.NAME, "address1").send_keys('Madhapur')
driver.find_element(By.NAME, "address2").send_keys('First Floor')
driver.find_element(By.NAME, "city").send_keys('Hyderabad')
select_state = Select(driver.find_element(By.XPATH, "//select[@id='id_state']"))
select_state.select_by_index(15)
driver.find_element(By.ID, "postcode").send_keys('54063')
select_country = Select(driver.find_element(By.XPATH, "//select[@id='id_country']"))
select_country.select_by_index(1)

driver.find_element(By.ID, "other").send_keys('This about my information')
driver.find_element(By.ID, "phone").send_keys('040-25648987')
driver.find_element(By.ID, "phone_mobile").send_keys('9894562574')
driver.find_element(By.ID, 'alias').send_keys('H No 1 2 102')
driver.find_element(By.ID, 'submitAccount').click()
time.sleep(3)
#
# path = 'C:/Users/FL_LPT-526/PycharmProjects/SecondAssessment/Information.xlsx'
# wb = openpyxl.load_workbook(path)
#
#
# # wb.create_sheet('Information')
# # data = [('Emai', 'Password'), (email, 'abcd1234')]
# # sheet = wb['Information']
# # for item in data:
# #     sheet.append(item)
# # wb.save(path)
#
# def set_emailid():
#     sheet = wb['Information']
#     email_id = sheet.cell(row=2, column=1)
#     email_id.value = email
#     wb.save(path)
#
#
# set_emailid()
# print('email', email)

# Login
# driver.get("http://automationpractice.com/")
# time.sleep(2)
# driver.find_element(By.XPATH, "//a[@class='login']").click()
# driver.find_element(By.ID, 'email').send_keys(email)
# driver.find_element(By.ID, 'passwd').send_keys('abcd1234')
# driver.find_element(By.ID, "SubmitLogin").click()

ele = driver.find_element(By.XPATH, '/html/body/div/div[1]/header/div[3]/div/div/div[6]/ul/li[2]/a')
time.sleep(2)
ActionChains(driver).move_to_element(ele).perform()
time.sleep(2)
driver.find_element(By.XPATH, "/html/body/div/div[1]/header/div[3]/div/div/div[6]/ul/li[2]/ul/li[1]/a").click()
# add to cart
product = driver.find_element(By.XPATH, "//div[@class='product-container']")
ActionChains(driver).move_to_element(product).perform()
driver.find_element(By.XPATH, "//a[@title='Add to cart']").click()

driver.find_element(By.XPATH, "//a[@title='Proceed to checkout']").click()
# dress_type = driver.find_element(By.XPATH, "//td[@class='cart_description'']").text
dress_type = driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[3]/div/div[2]/table/tbody/tr/td[2]/p/a").text
print('dress type is', dress_type)
total_price = driver.find_element(By.ID, "total_price").text
print('Total Price ', total_price)

driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[3]/div/p[2]/a[1]").click()

# wb.create_sheet('DressInformation')
# data = [('Dress Title', 'Price'), (dress_type, total_price)]
# sheet = wb['DressInformation']
# for item in data:
#     sheet.append(item)
# wb.save(path)
#
#
# #
# def set_Data():
#     sheet = wb['Information']
#     type = sheet.cell(row=2, column=1)
#     type.value = dress_type
#     price = sheet.cell(row=2, column=2)
#     type.value = price
#     wb.save(path)


# time.sleep(2)
# driver.find_element(By.ID, 'email').send_keys(random_email())
# time.sleep(2)
# driver.find_element(By.ID, 'passwd').send_keys('abcd1234')
# time.sleep(2)
# driver.find_element(By.ID, "SubmitLogin").click()

driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[3]/div/form/p/button").click()
time.sleep(2)
driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[3]/div/form/p/button").click()
time.sleep(2)
driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[3]/div/div/form/p/button").click()
time.sleep(2)
driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[3]/div/div/div[3]/div[1]/div/p/a").click()
time.sleep(2)
driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[3]/div/form/p/button/span").click()
time.sleep(2)
text = driver.find_element(By.XPATH, "/html/body/div/div[2]/div/div[3]/div/div/p").text
print(text)
pri = driver.find_element(By.XPATH, "//div[@class='box']").text

# wb.create_sheet('PaymentDetails')
# data = [('Amount', 'AccountOwner','Details','OrderReference'), ('Pradeep Macherla','Random','XPGAXQTVH')]
# sheet = wb['DressInformation']
#  for item in data:
#      sheet.append(item)
#  wb.save(path)
#
#
# #
#  def set_Data():
#     sheet = wb['Information']
#      type = sheet.cell(row=2, column=1)
#      type.value = dress_type
#      price = sheet.cell(row=2, column=2)
#      type.value = price
#      wb.save(path)
